import io
import os
from uuid import uuid4

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import login_required
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from app.decorators import feature_required, permission_required
from app.extensions import db
from app.models import AppSetting, AuditLog, Branch, ComplaintType, Employee, SystemFunction, User, UserFunctionAccess, UserPermission
from app.services.access import ROLE_ADMIN, ROLE_AGENT, ROLE_SUPERVISOR, apply_role_features
from app.services.audit import log_action

admin_bp = Blueprint("admin", __name__)

ALLOWED_IMAGE = {"png", "jpg", "jpeg", "gif", "webp"}


def _settings_dict():
    return {
        "brand_name": AppSetting.get("brand_name", "Lotus CRM"),
        "primary_color": AppSetting.get("primary_color", "#00796b"),
        "notification_email": AppSetting.get("notification_email", ""),
        "smtp_host": AppSetting.get("smtp_host", ""),
        "smtp_port": AppSetting.get("smtp_port", "587"),
        "smtp_user": AppSetting.get("smtp_user", ""),
        "smtp_password": AppSetting.get("smtp_password", ""),
        "use_graph_api": AppSetting.get("use_graph_api", "0"),
        "graph_tenant_id": AppSetting.get("graph_tenant_id", ""),
        "graph_client_id": AppSetting.get("graph_client_id", ""),
        "graph_client_secret": AppSetting.get("graph_client_secret", ""),
        "logo_path": AppSetting.get("logo_path", ""),
    }


@admin_bp.route("/")
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def index():
    users = User.query.order_by(User.username).all()
    functions = SystemFunction.query.order_by(SystemFunction.sort_order).all()
    access_map = {}
    for u in users:
        access_map[u.id] = {
            row.function_id: row.is_visible
            for row in UserFunctionAccess.query.filter_by(user_id=u.id).all()
        }
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.employee_name).all()
    branches = Branch.query.order_by(Branch.branch_name).all()
    complaint_types = ComplaintType.query.order_by(ComplaintType.sort_order).all()
    return render_template(
        "admin/index.html",
        users=users,
        functions=functions,
        access_map=access_map,
        employees=employees,
        branches=branches,
        complaint_types=complaint_types,
        settings=_settings_dict(),
        roles=[ROLE_AGENT, ROLE_SUPERVISOR, ROLE_ADMIN],
    )


@admin_bp.route("/users/add", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def add_user():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role", ROLE_AGENT)
    employee_code = request.form.get("employee_code") or None
    if not username or not password:
        flash("required_fields", "error")
        return redirect(url_for("admin.index"))
    if User.query.filter_by(username=username).first():
        flash("customer_exists", "error")
        return redirect(url_for("admin.index"))

    user = User(username=username, role=role, employee_code=employee_code)
    user.set_password(password)
    db.session.add(user)
    db.session.flush()

    perms = {
        ROLE_AGENT: dict(can_view_reports=False, can_edit_functions=False, can_manage_users=False, can_export_excel=False),
        ROLE_SUPERVISOR: dict(can_view_reports=True, can_edit_functions=False, can_manage_users=False, can_export_excel=True),
        ROLE_ADMIN: dict(can_view_reports=True, can_edit_functions=True, can_manage_users=True, can_export_excel=True),
    }
    p = perms.get(role, perms[ROLE_AGENT])
    db.session.add(UserPermission(user_id=user.id, **p))
    apply_role_features(user, role)
    db.session.commit()
    flash("user_created", "success")
    return redirect(url_for("admin.index"))


@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.username == "admin":
        flash("access_denied", "error")
        return redirect(url_for("admin.index"))
    db.session.delete(user)
    db.session.commit()
    flash("user_deleted", "success")
    return redirect(url_for("admin.index"))


@admin_bp.route("/users/<int:user_id>/role", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def update_user_role(user_id):
    user = User.query.get_or_404(user_id)
    role = request.form.get("role", ROLE_AGENT)
    user.role = role
    user.employee_code = request.form.get("employee_code") or None
    apply_role_features(user, role)
    db.session.commit()
    flash("permissions_updated", "success")
    return redirect(url_for("admin.index"))


@admin_bp.route("/permissions/<int:user_id>", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def update_permissions(user_id):
    perms = UserPermission.query.filter_by(user_id=user_id).first_or_404()
    perms.can_view_reports = "can_view_reports" in request.form
    perms.can_edit_functions = "can_edit_functions" in request.form
    perms.can_manage_users = "can_manage_users" in request.form
    perms.can_export_excel = "can_export_excel" in request.form
    db.session.commit()
    flash("permissions_updated", "success")
    return redirect(url_for("admin.index"))


@admin_bp.route("/functions/<int:function_id>/toggle", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_edit_functions")
def toggle_function(function_id):
    func = SystemFunction.query.get_or_404(function_id)
    func.is_enabled = not func.is_enabled
    db.session.commit()
    return redirect(url_for("admin.index", tab="features"))


@admin_bp.route("/access/set", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def set_access():
    user_id = int(request.form.get("user_id"))
    function_id = int(request.form.get("function_id"))
    visible = request.form.get("visible") == "1"
    access = UserFunctionAccess.query.filter_by(user_id=user_id, function_id=function_id).first()
    if access:
        access.is_visible = visible
    else:
        db.session.add(UserFunctionAccess(user_id=user_id, function_id=function_id, is_visible=visible))
    db.session.commit()
    flash("permissions_updated", "success")
    return redirect(url_for("admin.index", tab="access"))


@admin_bp.route("/settings", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def save_settings():
    keys = [
        "brand_name", "primary_color", "notification_email",
        "smtp_host", "smtp_port", "smtp_user", "smtp_password",
        "graph_tenant_id", "graph_client_id", "graph_client_secret",
    ]
    for key in keys:
        val = request.form.get(key, "")
        if val or key not in ("smtp_password", "graph_client_secret"):
            AppSetting.set(key, val)
    AppSetting.set("use_graph_api", "1" if request.form.get("use_graph_api") else "0")
    flash("settings_saved", "success")
    return redirect(url_for("admin.index", tab="settings"))


@admin_bp.route("/logo", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def upload_logo():
    file = request.files.get("logo")
    if not file or not file.filename:
        flash("required_fields", "error")
        return redirect(url_for("admin.index", tab="branding"))
    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ALLOWED_IMAGE:
        flash("access_denied", "error")
        return redirect(url_for("admin.index", tab="branding"))
    upload_dir = current_app.config["UPLOAD_FOLDER"]
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"logo.{ext}"
    path = os.path.join(upload_dir, filename)
    file.save(path)
    AppSetting.set("logo_path", f"uploads/{filename}")
    flash("logo_saved", "success")
    return redirect(url_for("admin.index", tab="branding"))


@admin_bp.route("/templates/branches")
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def template_branches():
    wb = Workbook()
    ws = wb.active
    ws.title = "Branches"
    ws.append(["branch_code", "branch_name", "branch_manager_email", "area_manager_email", "sales_manager_email"])
    ws.append(["B001", "Lotus Maadi", "manager@example.com", "area@example.com", "sales@example.com"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="branches_template.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route("/templates/employees")
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def template_employees():
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(["employee_code", "employee_name", "branch_code"])
    ws.append(["E001", "Ahmed Hassan", "B001"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="employees_template.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route("/import/branches", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def import_branches():
    file = request.files.get("file")
    if not file:
        flash("required_fields", "error")
        return redirect(url_for("admin.index", tab="import"))
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    count = 0
    for row in rows:
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        branch = Branch.query.get(code) or Branch(branch_code=code)
        branch.branch_name = str(row[1] or code).strip()
        branch.branch_manager_email = str(row[2] or "").strip() if len(row) > 2 else ""
        branch.area_manager_email = str(row[3] or "").strip() if len(row) > 3 else ""
        branch.sales_manager_email = str(row[4] or "").strip() if len(row) > 4 else ""
        db.session.add(branch)
        count += 1
    db.session.commit()
    flash("import_success", "success")
    return redirect(url_for("admin.index", tab="import"))


@admin_bp.route("/import/employees", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def import_employees():
    file = request.files.get("file")
    if not file:
        flash("required_fields", "error")
        return redirect(url_for("admin.index", tab="import"))
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        emp = Employee.query.get(code) or Employee(employee_code=code)
        emp.employee_name = str(row[1] or code).strip()
        emp.branch_code = str(row[2]).strip() if len(row) > 2 and row[2] else None
        emp.is_active = True
        db.session.add(emp)
    db.session.commit()
    flash("import_success", "success")
    return redirect(url_for("admin.index", tab="import"))


@admin_bp.route("/types/add", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def add_complaint_type():
    name_ar = request.form.get("name_ar", "").strip()
    name_en = request.form.get("name_en", "").strip()
    if not name_ar:
        flash("required_fields", "error")
        return redirect(url_for("admin.index", tab="types"))
    db.session.add(
        ComplaintType(
            name_ar=name_ar,
            name_en=name_en,
            requires_online="requires_online" in request.form,
            sort_order=ComplaintType.query.count() + 1,
        )
    )
    log_action("admin.type.add", "complaint_type", details=name_ar)
    db.session.commit()
    flash("updated", "success")
    return redirect(url_for("admin.index", tab="types"))


@admin_bp.route("/types/<int:type_id>/edit", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def edit_complaint_type(type_id):
    ct = ComplaintType.query.get_or_404(type_id)
    ct.name_ar = request.form.get("name_ar", ct.name_ar).strip()
    ct.name_en = request.form.get("name_en", ct.name_en or "").strip()
    ct.requires_online = "requires_online" in request.form
    ct.is_active = "is_active" in request.form
    log_action("admin.type.edit", "complaint_type", type_id, ct.name_ar)
    db.session.commit()
    flash("updated", "success")
    return redirect(url_for("admin.index", tab="types"))


@admin_bp.route("/branches/save", methods=["POST"])
@login_required
@feature_required("admin.index")
@permission_required("can_manage_users")
def save_branch_emails():
    code = request.form.get("branch_code")
    branch = Branch.query.get_or_404(code)
    branch.branch_name = request.form.get("branch_name", branch.branch_name).strip()
    branch.branch_manager_email = request.form.get("branch_manager_email", "").strip()
    branch.area_manager_email = request.form.get("area_manager_email", "").strip()
    branch.sales_manager_email = request.form.get("sales_manager_email", "").strip()
    log_action("admin.branch.edit", "branch", code)
    db.session.commit()
    flash("updated", "success")
    return redirect(url_for("admin.index", tab="branches"))


@admin_bp.route("/audit")
@login_required
@feature_required("admin.audit_logs")
@permission_required("can_manage_users")
def audit_logs():
    user_filter = request.args.get("user", "")
    q = AuditLog.query.order_by(AuditLog.created_at.desc())
    if user_filter:
        q = q.filter(AuditLog.username == user_filter)
    logs = q.limit(300).all()
    usernames = [u.username for u in User.query.order_by(User.username).all()]
    return render_template("admin/audit.html", logs=logs, usernames=usernames, user_filter=user_filter)
